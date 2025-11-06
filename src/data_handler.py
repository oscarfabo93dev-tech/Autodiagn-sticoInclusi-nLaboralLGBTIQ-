# src/data_handler.py
# -*- coding: utf-8 -*-
"""
Carga y parseo del Excel original:
- Instrucciones (sheet 'Instrucciones')
- Cuestionario (sheet 'Cuestionario'): preguntas A..J con opciones 3/2/1
- Niveles (sheets 'Nivel 1','Nivel 2','Nivel 3'): NIVEL, DEFINICION, CARACTERISTICAS, RUTA
- Recomendaciones (sheet 'Recomendaciones')
- Umbrales desde la fórmula en 'Cuestionario'!C81 si existe (fallback a 15/23)
"""

from __future__ import annotations
from pathlib import Path
from typing import Dict, Any, List, Optional
import re
import warnings

import pandas as pd
from openpyxl import load_workbook

# Suprimir warnings de openpyxl sobre extensiones no soportadas
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Nombre EXACTO del archivo esperado
EXCEL_FILENAME = (
    "Recurso 5.2. Autodiagnóstico en inclusión laboral LGBTIQ para empresas.xlsx"
)

DEFAULT_THRESHOLDS = {"nivel_1_max": 15, "nivel_2_max": 23}


def _norm_text(x: Any) -> str:
    return str(x).strip() if x is not None else ""


def _sheet_to_dataframe(path: Path, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet, dtype=str, engine="openpyxl")
    df = df.dropna(how="all").dropna(axis=1, how="all").fillna("")
    return df


def _looks_like_question(text: Any) -> bool:
    s = _norm_text(text)
    return ("¿" in s or "?" in s) and len(s) > 10


def _is_question_id(val: Any) -> bool:
    s = _norm_text(val)
    return len(s) == 1 and s.isalpha()


def _find_excel_path(data_dir: Path) -> Optional[Path]:
    xs = list(data_dir.glob("*.xlsx"))
    xs = [x for x in xs if not x.name.startswith("~$")]

    if not xs:
        return None
    if len(xs) == 1:
        return xs[0]

    for p in xs:
        name_lower = p.name.lower()
        if any(
            kw in name_lower for kw in ["autodiagnostico", "cuestionario", "lgbtiq"]
        ):
            return p

    return xs[0]


def _load_instructions_from_excel(xlsx: Path) -> str:
    df = _sheet_to_dataframe(xlsx, "Instrucciones")
    parts: List[str] = []
    for _, row in df.iterrows():
        for v in row.tolist():
            sv = _norm_text(v)
            if sv:
                parts.append(sv)

    seen, uniq = set(), []
    for p in parts:
        if p not in seen:
            uniq.append(p)
            seen.add(p)

    return (
        "\n".join(uniq)
        if uniq
        else "Bienvenido. Por favor lea y complete el cuestionario."
    )


def _load_questions_from_excel(xlsx: Path) -> List[Dict[str, Any]]:
    """OPTIMIZADO: Lee todo el sheet como DataFrame para ser 100x más rápido."""
    # Leer TODO el sheet de una vez (mucho más rápido)
    df = pd.read_excel(xlsx, sheet_name="Cuestionario", engine="openpyxl")

    # Asegurar que tenemos las columnas B, C, D (índices 1, 2, 3)
    if df.shape[1] < 4:
        raise ValueError(
            "El sheet 'Cuestionario' debe tener al menos 4 columnas (A, B, C, D)"
        )

    # Renombrar columnas para facilitar acceso
    df.columns = [f"col_{i}" for i in range(len(df.columns))]
    col_b = "col_1"  # Columna B
    col_c = "col_2"  # Columna C
    col_d = "col_3"  # Columna D

    questions: List[Dict[str, Any]] = []
    current_section: Optional[str] = None
    row_idx = 0
    max_rows = len(df)

    while row_idx < max_rows:
        b_val = df.iloc[row_idx][col_b]
        c_val = df.iloc[row_idx][col_c]

        # Detectar sección
        if (
            pd.notna(b_val)
            and not _is_question_id(b_val)
            and (pd.isna(c_val) or not _looks_like_question(c_val))
        ):
            sb = _norm_text(b_val)
            if sb and sb.lower() != "respuesta":
                current_section = sb

        # Detectar pregunta
        if _is_question_id(b_val) and _looks_like_question(c_val):
            qid = _norm_text(b_val)
            qtext = _norm_text(c_val)
            options: List[Dict[str, Any]] = []

            # Leer opciones (siguiente fila en adelante)
            r2 = row_idx + 1
            while r2 < max_rows:
                b2_val = df.iloc[r2][col_b]
                c2_val = df.iloc[r2][col_c]

                # Verificar si es un score válido (3, 2, o 1)
                if pd.notna(b2_val):
                    try:
                        score = int(float(str(b2_val)))
                        if (
                            score in (3, 2, 1)
                            and pd.notna(c2_val)
                            and _norm_text(c2_val)
                        ):
                            options.append(
                                {"score": score, "label": _norm_text(c2_val)}
                            )
                            r2 += 1
                            continue
                    except (ValueError, TypeError):
                        pass
                break

            # Completar opciones faltantes
            options = sorted(options, key=lambda x: x["score"], reverse=True)
            if len(options) != 3:
                missing = {3, 2, 1} - {o["score"] for o in options}
                for s in sorted(missing, reverse=True):
                    options.append({"score": s, "label": f"Opción {s}"})
                options = sorted(options, key=lambda x: x["score"], reverse=True)

            questions.append(
                {
                    "id": qid,
                    "section": current_section or "",
                    "text": qtext,
                    "options": options,
                }
            )
            row_idx = r2
            continue

        row_idx += 1

    if not questions:
        raise ValueError(
            "No se encontraron preguntas en 'Cuestionario'. Verifica el formato."
        )
    return questions


def _extract_thresholds_from_formula(xlsx: Path) -> Dict[str, int]:
    try:
        wb = load_workbook(xlsx, data_only=False, read_only=True)
        ws = wb["Cuestionario"]
        f = ws["C81"].value
        wb.close()

        if not isinstance(f, str) or "IF(" not in f.upper():
            return DEFAULT_THRESHOLDS.copy()
        comps = re.findall(r"D77\s*([<>=]{1,2})\s*(\d+)", f)
        nums = sorted({int(n) for (_op, n) in comps if n.isdigit()})
        if len(nums) >= 2:
            return {"nivel_1_max": nums[0], "nivel_2_max": nums[1]}
        if len(nums) == 1:
            return {
                "nivel_1_max": nums[0],
                "nivel_2_max": DEFAULT_THRESHOLDS["nivel_2_max"],
            }
        return DEFAULT_THRESHOLDS.copy()
    except Exception:
        return DEFAULT_THRESHOLDS.copy()


def _load_single_level_sheet(xlsx: Path, sheet_name: str) -> Dict[str, str]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb[sheet_name]
    out: Dict[str, str] = {}

    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c not in (None, "")]
        if len(cells) < 2:
            continue
        left = str(cells[0]).strip().upper()
        right = str(cells[1]).strip()
        mapping = {
            "NIVEL": "NIVEL",
            "DEFINICION": "DEFINICION",
            "DEFINICIÓN": "DEFINICION",
            "CARACTERISTICAS": "CARACTERISTICAS",
            "CARACTERÍSTICAS": "CARACTERISTICAS",
            "RUTA": "RUTA",
            "RUTA DE APRENDIZAJE SUGERIDA": "RUTA",
            "RUTA_DE_APRENDIZAJE_SUGERIDA": "RUTA",
        }
        key = mapping.get(left)
        if key and right:
            out[key] = f"{out[key]}\n{right}" if key in out else right

    wb.close()

    if not out:
        df = _sheet_to_dataframe(xlsx, sheet_name)
        for col in df.columns:
            cu = str(col).strip().upper()
            if cu in [
                "NIVEL",
                "DEFINICION",
                "DEFINICIÓN",
                "CARACTERISTICAS",
                "CARACTERÍSTICAS",
                "RUTA",
                "RUTA_DE_APRENDIZAJE_SUGERIDA",
            ]:
                text = "\n".join(
                    [_norm_text(x) for x in df[col].tolist() if _norm_text(x)]
                )
                cu = "DEFINICION" if cu == "DEFINICIÓN" else cu
                cu = "CARACTERISTICAS" if cu == "CARACTERÍSTICAS" else cu
                out[cu] = text
    return out


def _load_levels_from_excel(xlsx: Path) -> Dict[str, Dict[str, str]]:
    out = {}
    for sn in ["Nivel 1", "Nivel 2", "Nivel 3"]:
        try:
            out[sn] = _load_single_level_sheet(xlsx, sn)
        except Exception:
            out[sn] = {}
    return out


def _normalize_rec_columns(cols: List[str]) -> List[str]:
    normed = []
    for c in cols:
        s = _norm_text(c).lower()
        s = (
            s.replace("á", "a")
            .replace("é", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ú", "u")
            .replace("ñ", "n")
        )
        s = s.replace(" ", "_")
        normed.append(s)
    return normed


def _load_recommendations_from_excel(xlsx: Path) -> pd.DataFrame:
    df = _sheet_to_dataframe(xlsx, "Recomendaciones")
    df.columns = _normalize_rec_columns(df.columns.tolist())
    mapping = {}
    for c in df.columns:
        if "barrera" in c:
            mapping[c] = "barrera"
        elif "concepto" in c:
            mapping[c] = "concepto"
        elif "sintoma" in c:
            mapping[c] = "sintomas"
        elif "indicador" in c:
            mapping[c] = "indicadores"
        elif "recomendacion" in c:
            mapping[c] = "recomendaciones"
    if mapping:
        df = df.rename(columns=mapping)
    return df


def load_data_from_excel(excel_path: str | Path) -> Dict[str, Any]:
    """Carga datos del Excel."""
    xlsx = Path(excel_path)
    if not xlsx.exists():
        raise FileNotFoundError(f"No encontré el Excel en: {xlsx}")

    instructions = _load_instructions_from_excel(xlsx)
    questions = _load_questions_from_excel(xlsx)
    thresholds = _extract_thresholds_from_formula(xlsx)
    levels = _load_levels_from_excel(xlsx)
    recs = _load_recommendations_from_excel(xlsx)

    return {
        "instructions": instructions,
        "questions": questions,
        "thresholds": thresholds,
        "levels": levels,
        "recommendations": recs,
    }


def load_data(data_dir: str | Path = "data") -> Dict[str, Any]:
    """Carga datos desde la carpeta data/ (sin debug logging)."""
    d = Path(data_dir)

    if not d.exists():
        raise FileNotFoundError(f"La carpeta {d.absolute()} no existe")

    xlsx = _find_excel_path(d)

    if not xlsx:
        available = list(d.glob("*.xlsx"))
        raise FileNotFoundError(
            f"No encontré ningún archivo Excel en {d.absolute()}. "
            f"Archivos disponibles: {[f.name for f in available]}"
        )

    return load_data_from_excel(xlsx)
